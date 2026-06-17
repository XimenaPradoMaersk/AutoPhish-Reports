"""
Data Preprocessing Module

This script replicates the logic implemented in a Dify code node.
It processes phishing simulation data by:
- Reading an Excel file
- Removing duplicate users based on email
- Prioritizing user actions (click > open > deliver > bounce)
- Outputting a clean dataset for further analysis
"""

from io import BytesIO
from openpyxl import load_workbook


def load_excel(file_path):
    """Load Excel file and return worksheet"""
    workbook = load_workbook(file_path, data_only=True)
    return workbook.active


def detect_columns(header):
    """Detect email and response columns dynamically"""
    email_index = None
    response_index = None

    for i, col_name in enumerate(header):
        col_name_str = str(col_name).strip().lower()

        if 'email' in col_name_str:
            email_index = i
        if 'response' in col_name_str:
            response_index = i

    return email_index, response_index


def get_priority(response):
    """Assign priority based on user action"""
    response = str(response).strip().lower()

    if 'click' in response:
        return 3
    elif 'open' in response:
        return 2
    elif 'deliver' in response:
        return 1
    elif 'bounce' in response:
        return 0
    else:
        return 0


def remove_duplicates(worksheet):
    """Deduplicate users keeping highest-risk action"""

    # Read header
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

    email_index, response_index = detect_columns(header)

    if email_index is None:
        raise ValueError("Email column not found")

    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    original_rows = len(rows)

    best_rows = {}

    for row in rows:
        email = str(row[email_index]).strip().lower()

        if not email or email == 'none':
            continue

        response = row[response_index] if response_index is not None else ''
        current_priority = get_priority(response)

        if email not in best_rows:
            best_rows[email] = (current_priority, row)
        else:
            if current_priority > best_rows[email][0]:
                best_rows[email] = (current_priority, row)

    # Build cleaned dataset
    output_data = []

    for _, (_, row) in best_rows.items():
        row_dict = {}
        for i, header_name in enumerate(header):
            row_dict[header_name] = row[i]
        output_data.append(row_dict)

    return {
        "data": output_data,
        "original_rows": original_rows,
        "final_rows": len(output_data),
        "duplicates_removed": original_rows - len(output_data)
    }


def main(file_path):
    worksheet = load_excel(file_path)
    result = remove_duplicates(worksheet)

    print("✅ Original rows:", result["original_rows"])
    print("✅ Final rows:", result["final_rows"])
    print("✅ Duplicates removed:", result["duplicates_removed"])

    return result


if __name__ == "__main__":
    # Example usage (replace with your file path)
    file_path = "sample_phishing_data.xlsx"
    main(file_path)
